import tkinter as tk
import threading
from tkinter import ttk, messagebox, PhotoImage, Button, Entry, filedialog
from pathlib import Path
from PIL import Image, ImageTk
import serial.tools.list_ports
import serial
import os
import json
import math
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

# Dirección de carpeta que contiene las imagenes de los botones
_DIR = os.path.dirname(__file__)
OUTPUT_PATH = Path(__file__).resolve().parent
ASSETS_PATH = OUTPUT_PATH / "assets" / "frame0"


# Evita que se ingresen valores no númericos a la entrada de edad
def validate_number_input(new_value):
    if new_value == "":
        return True
    try:
        float(new_value)
        return True
    except ValueError:
        return False


# Permite llamar la imagen de un botón
def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / Path(path)  # La entrada es el nombre de la imagen como variable str


def read_input_from_json(variable, filename="user_input.json"):
    try:
        with open(filename, "r") as file:
            data = json.load(file)
            return data.get(variable)  # Using .get() to avoid KeyError
    except (FileNotFoundError, json.JSONDecodeError):
        return None  # File doesn't exist or is invalid


"""
########################################################################################################################

La clase de "Controller" tiene la tarea de realizar las sig. funciones
    * Permite el cambio entre ventanas de la aplicación
    * Declara las acciones de seguridad necesarias al cambiar y cerrar ventanas
    * Declara las variables de inf. del usuario que se comparten con las demas ventanas
    
########################################################################################################################
"""


class Controller:
    def __init__(self, root):
        self.root = root
        self.patient_data = {}  # Crea un diccionario vacío en el que se almacena la inf. del paciente
        self.config_data = {}
        self.current_frame = None
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)  # Enlaza el evento on_closing al protoc. de ventana

    """Cambio entre ventanas."""

    def switch_frame(self, new_frame_class):
        if hasattr(self.current_frame, 'disconnect_arduino'):
            self.current_frame.disconnect_arduino()  # Manda a llamar la función de desconexión dentro de cada ventana
        if self.current_frame:
            self.current_frame.pack_forget()  # Oculta la ventana actual
        self.current_frame = new_frame_class(self.root, self, self.patient_data, self.config_data)
        self.current_frame.pack(fill="both", expand=True)

    """Controla el evento al cerrar ventana."""

    def on_closing(self):
        if hasattr(self.current_frame, 'turn_off_motor'):
            self.current_frame.turn_off_motor()  # Encuentra y activa función de apagado del motor
        if hasattr(self.current_frame, 'disconnect_arduino'):
            self.current_frame.disconnect_arduino()  # Encuentra y activa la función que desconecta del arduino
        time.sleep(0.05)
        self.root.destroy()  # Cierra la ventana


"""
########################################################################################################################
La clase de "AppBase" contiene el formato y colores que la interfaz necesita

########################################################################################################################
"""


class AppBase(tk.Frame):
    def __init__(self, root, controller, patient_data, config_data):
        super().__init__(root)
        self.root = root
        self.controller = controller
        self.patient_data = patient_data  # Shared patient data
        self.config_data = config_data
        self.used_color = 'white'
        self.text_color = "#3c3d40"
        self.connected_color = "#00487C"
        self.disconnected_color = "#9E2B25"
        self.root.configure(bg=self.used_color)
        self.validate_func = self.register(validate_number_input)  # Validación numérica


"""
############################################ PRIMERA VENTANA ###########################################################
Se reciben los datos del paciente
    * Ingresa: nombre, # de expediente, edad, sexo. act. física y fecha
    * Establece la dirección del archivo a generar con los datos del estudio

########################################################################################################################
"""


class AppInterface0(AppBase):
    def __init__(self, root, controller, patient_data, config_data):
        super().__init__(root, controller, patient_data, config_data)
        self.pack(fill="both", expand=True)
        self.dimension_x0 = "630"
        self.dimension_y0 = "550"

        self.screen_width = self.root.winfo_screenwidth()  # Obtiene el ancho de la pantalla
        self.screen_height = self.root.winfo_screenheight()  # Obtiene el alto de la pantalla
        self.x0 = (self.screen_width // 2) - (int(self.dimension_x0) // 2)  # Calcula la posición X
        self.y0 = (self.screen_height // 2) - (int(self.dimension_y0) // 2)  # Calcula la posición Y
        self.root.geometry(f"{self.dimension_x0}x{self.dimension_y0}+{self.x0}+{self.y0}")
        self.root.title("ITESM ETI Prototype")

        self.root.resizable(False, False)
        self.root.configure(bg=self.used_color)
        self.canvas = self.create_canvas(self.dimension_y0, self.dimension_x0)
        self.validate_func = self.canvas.register(validate_number_input)
        self.init_widgets()

    """función que crea el lienzo en el que se colocan los objetos dentro de la ventana"""

    def create_canvas(self, dimension_y, dimension_x):
        canvas = tk.Canvas(
            self.root,
            bg=self.used_color,
            height=int(dimension_y),
            width=int(dimension_x),
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )
        canvas.place(x=0, y=0)
        return canvas

    """Botones de la página número 0"""

    def init_widgets(self):
        # Titulo
        self.canvas.create_text(140, 40, anchor="w", text="Registro de paciente", fill=self.text_color,
                                font=("Inter", 30), width=400)

        # Nombre del paciente
        self.patient_bg_image = PhotoImage(file=relative_to_assets("PATIENT_ENTR_BG.png"))
        self.canvas.create_image(105, 101, image=self.patient_bg_image, anchor="nw")
        self.entry_00 = Entry(bd=0, bg="white", fg="#000000", highlightthickness=0, font=("Inter", 13))
        self.entry_00.place(x=140.0, y=127.0, width=350.0, height=23.0)

        # Numero de expediente
        self.exp_bg_image = PhotoImage(file=relative_to_assets("EXP_ENTR_BG.png"))
        self.canvas.create_image(105, 169, image=self.exp_bg_image, anchor="nw")
        self.entry_02 = Entry(bd=0, bg="white", fg="#000000", highlightthickness=0, state="normal", font=("Inter", 13))
        self.entry_02.place(x=140.0, y=195.0, width=320.0, height=20.0)

        # Edad
        self.age_bg_image = PhotoImage(file=relative_to_assets("AGE_ENTRY_BG.png"))
        self.canvas.create_image(105, 237, image=self.age_bg_image, anchor="nw")
        self.entry_01 = Entry(bd=0, bg="white", fg="#000000", highlightthickness=0, state="normal",
                              font=("Inter", 13), validate="key", validatecommand=(self.validate_func, "%P"))
        self.entry_01.place(x=140.0, y=263.0, width=100.0, height=20.0)

        # Sexo
        self.sex_bg_image = PhotoImage(file=relative_to_assets("SEX_ENTRY_BG.png"))
        self.canvas.create_image(327, 237, image=self.sex_bg_image, anchor="nw")
        genders = ["Masculino", "Femenino"]
        self.combobox1 = ttk.Combobox(self.canvas, values=genders, font=("Inter", 12))
        self.combobox1.config(state="readonly")
        self.combobox1.set("...")
        self.combobox1.place(x=360.0, y=262.0, width=150.0, height=24)

        # Actividad física
        self.act_bg_image = PhotoImage(file=relative_to_assets("ACT_ENTRY_BG.png"))
        self.canvas.create_image(105, 308, image=self.act_bg_image, anchor="nw")
        lifestl = ["Sedentario", "Actividad moderada", "Deportista"]
        self.combobox2 = ttk.Combobox(self.canvas, values=lifestl, font=("Inter", 13))
        self.combobox2.config(state="readonly")
        self.combobox2.set("...")
        self.combobox2.place(x=140.0, y=333.0, width=250.0, height=25)

        # Fecha del estudio
        self.date_bg_image = PhotoImage(file=relative_to_assets("DATE_ENTRY_BG.png"))
        self.canvas.create_image(105, 376, image=self.date_bg_image, anchor="nw")
        self.day_label = ttk.Label(root, text="Day:")
        self.day_combobox = ttk.Combobox(root, values=[str(i).zfill(2) for i in range(1, 32)])
        self.day_combobox.set("01")
        self.day_combobox.place(x=120, y=397, width=50)
        self.month_label = ttk.Label(root, text="Month:")
        self.month_combobox = ttk.Combobox(root, values=[str(i).zfill(2) for i in range(1, 13)])
        self.month_combobox.set("01")
        self.month_combobox.place(x=177, y=397, width=50)
        self.year_label = ttk.Label(root, text="Year:")
        self.year_combobox = ttk.Combobox(root, values=[str(i) for i in range(2000, 2031)])
        self.year_combobox.set("2023")
        self.year_combobox.place(x=235, y=397, width=60)

        # Botón de cambio de pantalla y almacenamiento de inf. en variables
        self.register_bg_image = PhotoImage(file=relative_to_assets("SWITCH_BTN_BG.png"))
        self.switch_button = tk.Button(self.canvas, image=self.register_bg_image, text="Go to Interface 1",
                                       command=self.save_and_next, state="normal", relief="flat",
                                       borderwidth=0, bg=self.used_color, )
        self.switch_button.place(x=224.0, y=460.0)

        # Botón a página de ajustes
        self.settings_bg_image = PhotoImage(file=relative_to_assets("COG_BG.png"))
        self.settings_button = Button(
            self.canvas,
            image=self.settings_bg_image,
            relief="flat",
            bg=self.used_color,
            command=lambda: self.controller.switch_frame(AppInterface01)

        )
        self.settings_button.place(x=544, y=460)

        # Selección de dirección del archivo
        self.select_bg_image = PhotoImage(file=relative_to_assets("SELECT_BG_IMAGE.png"))
        self.select_folder_button = Button(
            self.canvas,
            image=self.select_bg_image,
            command=self.select_output_folder,
            relief="flat",
            state="normal",
            bg=self.used_color
        )
        self.select_folder_button.place(x=385, y=386)

    def select_output_folder(self):
        """función que asigna la dirección del archivo"""
        folder_selected = filedialog.askdirectory()  # Abre un diálogo para seleccionar la carpeta de salida.
        if folder_selected:
            self.output_folder = Path(folder_selected)
            self.patient_data["output_folder"] = self.output_folder  # Guardar la ruta en patient_data
            messagebox.showinfo("Carpeta seleccionada", f"Los archivos se guardarán en: {self.output_folder}")

    def save_and_next(self):
        """Guarda la información y cambia de interfaz."""
        self.patient_data["Nombre"] = self.entry_00.get()
        self.patient_data["Edad"] = self.entry_01.get()
        self.patient_data["Sexo"] = self.combobox1.get()
        self.patient_data["Actividad"] = self.combobox2.get()
        self.patient_data["Expediente"] = self.entry_02.get()
        self.patient_data["Fecha"] = f"{self.day_combobox.get()}/{self.month_combobox.get()}/{self.year_combobox.get()}"


        if not all(self.patient_data.values()):
            messagebox.showwarning("Error", "Asegúrese de llenar todos los espacios")
            return

        self.controller.switch_frame(AppInterface1)

    def error_message(self):
        messagebox.showwarning("Error", "Asegurese de llenar todos los espacios")


"""
############################################ SEGUNDA VENTANA ###########################################################
Presenta las posibles configuraciones

########################################################################################################################
"""


class AppInterface01(AppBase):
    def __init__(self, root, controller, patient_data, config_data):
        super().__init__(root, controller, patient_data, config_data)
        self.pack(fill="both", expand=True)
        self.dimension_x0 = "850"
        self.dimension_y0 = "550"

        self.screen_width = self.root.winfo_screenwidth()  # Obtiene el ancho de la pantalla
        self.screen_height = self.root.winfo_screenheight()  # Obtiene el alto de la pantalla
        self.x0 = (self.screen_width // 2) - (int(self.dimension_x0) // 2)  # Calcula la posición X
        self.y0 = (self.screen_height // 2) - (int(self.dimension_y0) // 2)  # Calcula la posición Y
        self.root.geometry(f"{self.dimension_x0}x{self.dimension_y0}+{self.x0}+{self.y0}")

        self.root.resizable(False, False)
        self.root.configure(bg=self.used_color)
        self.canvas = self.create_canvas(self.dimension_y0, self.dimension_x0)
        self.validate_func = self.canvas.register(validate_number_input)
        self.init_widgets()
        self.user_input_widgets()
        self.root.title("ITESM ETI Prototype")

    def create_canvas(self, dimension_y, dimension_x):
        canvas = tk.Canvas(self.root, bg=self.used_color, height=int(dimension_y), width=int(dimension_x), bd=0,
                           highlightthickness=0, relief="ridge")
        canvas.place(x=0, y=0)
        return canvas

    def save_input_to_json(self, value, variable, filename="user_input.json"):
        # Load existing data if file exists
        data = {}
        if os.path.exists(filename):
            try:
                with open(filename, "r") as file:
                    data = json.load(file)
            except (FileNotFoundError, json.JSONDecodeError):
                data = {}  # If file is corrupt, start fresh

        # Update with new data
        data[variable] = value

        # Write back to file
        with open(filename, "w") as file:
            json.dump(data, file, indent=3)  # indent for pretty formatting

    def init_widgets(self):
        self.tit_st = tk.Label(self.canvas, text="Configuración", fg=self.text_color, font=("Inter", 30),
                               bg=self.used_color)
        self.tit_st.place(x=300.0, y=10.0, width=250)

        self.go_back_bg_image = PhotoImage(file=relative_to_assets("GO_BACK_BTN1.png"))
        self.go_back_button = tk.Button(self.canvas, image=self.go_back_bg_image,
                                        command=lambda: self.controller.switch_frame(AppInterface0), state="normal",
                                        relief="flat", borderwidth=0, bg=self.used_color)
        self.go_back_button.place(x=755.0, y=465.0)

    def read_input_from_json(self, variable, filename="user_input.json"):
        try:
            with open(filename, "r") as file:
                data = json.load(file)
                return data.get(variable)  # Using .get() to avoid KeyError
        except (FileNotFoundError, json.JSONDecodeError):
            return None  # File doesn't exist or is invalid

    def save_config(self):
        set1 = [self.sed_nv1_entry, self.sed_nv2_entry, self.sed_nv3_entry,
                self.sed_nv4_entry, self.sed_nv5_entry, self.sed_time_entry]
        set2 = [self.mod_nv1_entry, self.mod_nv2_entry, self.mod_nv3_entry,
                self.mod_nv4_entry, self.mod_nv5_entry, self.mod_time_entry]
        set3 = [self.dep_nv1_entry, self.dep_nv2_entry, self.dep_nv3_entry,
                self.dep_nv4_entry, self.dep_nv5_entry, self.dep_time_entry]

        set_matrix = [set1, set2, set3]
        set_names = ["sed", "mod", "dep"]  # (JSON keys)

        for set_name, row in zip(set_names, set_matrix):
            for i, entry in enumerate(row, start=1):
                value = entry.get().strip()  # Get and trim whitespace
                key = f"{set_name}_nv{i}" if i <= 5 else f"{set_name}_time"  # Dynamic key naming

                if not value:
                    self.conform = self.read_input_from_json(key)
                    self.save_input_to_json(self.conform, key)
                elif int(value) >= 18:
                    messagebox.showwarning("Error", "La potencia nominal del motor es 18 N/m, \n"
                                                    "ingrese un valor menor")
                else:
                    self.save_input_to_json(value, key)

        # Borra las entradas al terminar de almacenar los datos
        for entry_set in [set1, set2, set3]:
            for entry in entry_set:
                entry.delete(0, "end")

        configs = [
            ('snv', 'sed_nv', 5), ('snvt', 'sed_time', 1),
            ('mnv', 'mod_nv', 5), ('mnt', 'mod_time', 1),
            ('pnv', 'dep_nv', 5), ('pnt', 'dep_time', 1)
        ]

        for prefix, json_prefix, count in configs:
            for i in range(1, count + 1):
                attr = f"{prefix}{i if count > 1 else ''}"
                json_key = f"{json_prefix}{i if count > 1 else ''}"
                getattr(self, attr).config(text=f"{read_input_from_json(json_key)}", fg=self.text_color, bg="#FFFFFF")

    def user_input_widgets(self):

        # Apartado de configuraciones para actividad sedentaria
        self.sed_bg_image = PhotoImage(file=relative_to_assets("SED_BG.png"))
        self.canvas.create_image(18, 78, image=self.sed_bg_image, anchor="nw")
        self.text_sed_bg_image = PhotoImage(file=relative_to_assets("TEXT_SET_BG.png"))
        self.canvas.create_image(40, 120, image=self.text_sed_bg_image, anchor="nw")

        self.snv1 = tk.Label(self.canvas, text=f"{read_input_from_json('sed_nv1')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.snv1.place(x=150, y=140, width=30)
        self.sed_nv1_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.sed_nv1_entry.place(x=215.0, y=140.0, width=50.0, height=20.0)

        self.snv2 = tk.Label(self.canvas, text=f"{read_input_from_json('sed_nv2')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.snv2.place(x=150, y=178, width=30)
        self.sed_nv2_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.sed_nv2_entry.place(x=215.0, y=178.0, width=50.0, height=20.0)

        self.snv3 = tk.Label(self.canvas, text=f"{read_input_from_json('sed_nv3')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.snv3.place(x=150, y=220, width=30)
        self.sed_nv3_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.sed_nv3_entry.place(x=215, y=217.0, width=50.0, height=20.0)

        self.snv4 = tk.Label(self.canvas, text=f"{read_input_from_json('sed_nv4')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.snv4.place(x=150, y=269, width=30)
        self.sed_nv4_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.sed_nv4_entry.place(x=215.0, y=264.0, width=50.0, height=20.0)

        self.snv5 = tk.Label(self.canvas, text=f"{read_input_from_json('sed_nv5')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.snv5.place(x=150, y=305, width=30)
        self.sed_nv5_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.sed_nv5_entry.place(x=215.0, y=303.0, width=50.0, height=20.0)

        self.snvt = tk.Label(self.canvas, text=f"{read_input_from_json('sed_time')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.snvt.place(x=150, y=389, width=30)
        self.sed_time_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                    font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.sed_time_entry.place(x=215.0, y=387.0, width=50.0, height=20.0)

        # Apartado de configuraciones para actividad moderada
        self.mod_bg_image = PhotoImage(file=relative_to_assets("MOD_BG.png"))
        self.canvas.create_image(298, 78, image=self.mod_bg_image, anchor="nw")
        self.text_mod_bg_image = PhotoImage(file=relative_to_assets("TEXT_SET_BG.png"))
        self.canvas.create_image(317, 120, image=self.text_sed_bg_image, anchor="nw")

        self.mnv1 = tk.Label(self.canvas, text=f"{read_input_from_json('mod_nv1')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.mnv1.place(x=427, y=140, width=30)
        self.mod_nv1_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.mod_nv1_entry.place(x=492.0, y=140.0, width=50.0, height=20.0)

        self.mnv2 = tk.Label(self.canvas, text=f"{read_input_from_json('mod_nv2')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.mnv2.place(x=427, y=178, width=30)
        self.mod_nv2_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.mod_nv2_entry.place(x=492.0, y=178.0, width=50.0, height=20.0)

        self.mnv3 = tk.Label(self.canvas, text=f"{read_input_from_json('mod_nv3')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.mnv3.place(x=427, y=220, width=30)
        self.mod_nv3_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.mod_nv3_entry.place(x=492.0, y=217.0, width=50.0, height=20.0)

        self.mnv4 = tk.Label(self.canvas, text=f"{read_input_from_json('mod_nv4')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.mnv4.place(x=427, y=269, width=30)
        self.mod_nv4_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.mod_nv4_entry.place(x=492.0, y=264.0, width=50.0, height=20.0)

        self.mnv5 = tk.Label(self.canvas, text=f"{read_input_from_json('mod_nv5')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.mnv5.place(x=427, y=305, width=30)
        self.mod_nv5_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.mod_nv5_entry.place(x=492.0, y=303.0, width=50.0, height=20.0)

        self.mnt = tk.Label(self.canvas, text=f"{read_input_from_json('mod_time')}", fg=self.text_color,
                            font=("Inter", 11), justify="center", bg=self.used_color)
        self.mnt.place(x=427, y=389, width=30)
        self.mod_time_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                    font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.mod_time_entry.place(x=492.0, y=387.0, width=50.0, height=20.0)

        # Apartado de configuraciones para actividad deportiva
        self.dep_bg_image = PhotoImage(file=relative_to_assets("SPOR_BG.png"))
        self.canvas.create_image(578, 77, image=self.dep_bg_image, anchor="nw")
        self.text_dep_bg_image = PhotoImage(file=relative_to_assets("TEXT_SET_BG.png"))
        self.canvas.create_image(595, 120, image=self.text_dep_bg_image, anchor="nw")

        self.pnv1 = tk.Label(self.canvas, text=f"{read_input_from_json('dep_nv1')}", fg=self.text_color,
                             font=("Inter", 11), justify="center",
                             bg=self.used_color)
        self.pnv1.place(x=708, y=143, width=30)
        self.dep_nv1_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.dep_nv1_entry.place(x=769.0, y=140.0, width=50.0, height=20.0)

        self.pnv2 = tk.Label(self.canvas, text=f"{read_input_from_json('dep_nv2')}", fg=self.text_color,
                             font=("Inter", 11), justify="center",
                             bg=self.used_color)
        self.pnv2.place(x=708, y=178, width=30)
        self.dep_nv2_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.dep_nv2_entry.place(x=769.0, y=178.0, width=50.0, height=20.0)

        self.pnv3 = tk.Label(self.canvas, text=f"{read_input_from_json('dep_nv3')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.pnv3.place(x=708, y=217, width=30)
        self.dep_nv3_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.dep_nv3_entry.place(x=769.0, y=217.0, width=50.0, height=20.0)

        self.pnv4 = tk.Label(self.canvas, text=f"{read_input_from_json('dep_nv4')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.pnv4.place(x=708, y=266, width=30)
        self.dep_nv4_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.dep_nv4_entry.place(x=769.0, y=264.0, width=50.0, height=20.0)

        self.pnv5 = tk.Label(self.canvas, text=f"{read_input_from_json('dep_nv5')}", fg=self.text_color,
                             font=("Inter", 11), justify="center", bg=self.used_color)
        self.pnv5.place(x=708, y=304, width=30)
        self.dep_nv5_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                   font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.dep_nv5_entry.place(x=769.0, y=303.0, width=50.0, height=20.0)

        self.pnt = tk.Label(self.canvas, text=f"{read_input_from_json('dep_time')}", fg=self.text_color,
                            font=("Inter", 11), justify="center", bg=self.used_color)
        self.pnt.place(x=708, y=387, width=30)
        self.dep_time_entry = Entry(bd=1, bg="white", fg="#000000", highlightthickness=0, state="normal",
                                    font=("Inter", 12), validate="key", validatecommand=(self.validate_func, "%P"))
        self.dep_time_entry.place(x=769.0, y=387.0, width=50.0, height=20.0)

        self.register_bg_image = PhotoImage(file=relative_to_assets("SAVE_ANGLE_BTN0_BG.png"))
        self.save_config_button = tk.Button(self.canvas, image=self.register_bg_image, text="Go to Interface 1",
                                            command=self.save_config, state="normal", relief="flat",
                                            borderwidth=0, bg=self.used_color, )
        self.save_config_button.place(x=325.0, y=460.0)


"""
############################################## TERCERA VENTANA #########################################################
Presenta las modalidades: AUTOMÁTICO Y MANUAL

########################################################################################################################
"""


class AppInterface1(AppBase):
    def __init__(self, root, controller, patient_data, config_data):
        super().__init__(root, controller, patient_data, config_data)
        self.pack(fill="both", expand=True)
        self.root = root
        self.dimension_x0 = "630"
        self.dimension_y0 = "550"
        self.root.resizable(False, False)
        self.root.configure(bg=self.used_color)
        self.root.title("ITESM ETI Prototype")

        self.screen_width = self.root.winfo_screenwidth()  # Obtiene el ancho de la pantalla
        self.screen_height = self.root.winfo_screenheight()  # Obtiene el alto de la pantalla
        self.x0 = (self.screen_width // 2) - (int(self.dimension_x0) // 2)  # Calcula la posición X
        self.y0 = (self.screen_height // 2) - (int(self.dimension_y0) // 2)  # Calcula la posición Y
        self.root.geometry(f"{self.dimension_x0}x{self.dimension_y0}+{self.x0}+{self.y0}")

        self.canvas = self.create_canvas(self.dimension_y0, self.dimension_x0)
        self.init_widgets()

    def create_canvas(self, dimension_y, dimension_x):
        canvas = tk.Canvas(
            self.root,
            bg=self.used_color,
            height=int(dimension_y),
            width=int(dimension_x),
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )
        canvas.place(x=0, y=0)
        return canvas

    def init_widgets(self):
        self.canvas.create_text(130, 40, anchor="w", text="Seleccione modalidad", fill=self.text_color,
                                font=("Inter", 30), width=400)
        self.canvas.create_text(135, 370, anchor="nw", text="MANUAL", fill=self.text_color, font=("Inter", 14))
        self.manual_bg_image = PhotoImage(file=relative_to_assets("MANUAL_BTN_BG.png"))
        self.manual_button = tk.Button(self.canvas, image=self.manual_bg_image,
                                       command=lambda: self.controller.switch_frame(AppInterface3), state="normal",
                                       relief="flat",
                                       borderwidth=0, bg=self.used_color, )
        self.manual_button.place(x=97.0, y=198.0)

        self.canvas.create_text(405, 370, anchor="nw", text="AUTOMÁTICO", fill=self.text_color, font=("Inter", 14))
        self.automatic_bg_image = PhotoImage(file=relative_to_assets("AUTOMATIC_BTN.png"))
        self.automatic_button = tk.Button(self.canvas, image=self.automatic_bg_image,
                                          command=lambda: self.controller.switch_frame(AppInterface2), state="normal",
                                          relief="flat", borderwidth=0, bg=self.used_color)
        self.automatic_button.place(x=381.0, y=183.0)

        self.go_back_bg_image = PhotoImage(file=relative_to_assets("GO_BACK_BTN1.png"))
        self.go_back_button = tk.Button(self.canvas, image=self.go_back_bg_image,
                                        command=lambda: self.controller.switch_frame(AppInterface0), state="normal",
                                        relief="flat", borderwidth=0, bg=self.used_color)
        self.go_back_button.place(x=520.0, y=474.0)


"""
############################################## CUARTA VENTANA #########################################################
Ventana que realiza el estudio AUTOMÁTICO

########################################################################################################################
"""


class AppInterface2(AppBase):
    def __init__(self, root, controller, patient_data, config_data):
        super().__init__(root, controller, patient_data, config_data)
        self.columns = None
        self.column_dimensions = None
        self.position = None
        self.messagebox = None
        self.level = None
        self.value = None
        self.text = None
        self.combobox = None
        self.connect_button_image = None
        self.leg_animation = None
        self.ser = None
        self.send_value_running = False
        self.send_value_thread = None
        self.max_torque_reached = False
        self.stop_threads = False
        self.is_connected = False
        self.message_shown = False
        self.active_animation = True
        self.blink_state = True
        self.grados = 0
        self.torque = 0
        self.nivel_actual = None
        self.datos = {
            "Nivel 1": [],
            "Nivel 2": [],
            "Nivel 3": [],
            "Nivel 4": [],
            "Nivel 5": []
        }
        self.time_left = 0
        self.max_angle = 0
        self.running = False
        self.frames_path = Path(__file__).resolve().parent / "light_video"
        self.squares = []
        self.root = root
        self.dimension_x1 = "1000"
        self.dimension_y1 = "720"

        self.screen_width = self.root.winfo_screenwidth()  # Obtiene el ancho de la pantalla
        self.screen_height = self.root.winfo_screenheight()  # Obtiene el alto de la pantalla
        x = (self.screen_width // 2) - (int(self.dimension_x1) // 2)  # Calcula la posición X
        y = (self.screen_height // 2) - (int(self.dimension_y1) // 2)  # Calcula la posición Y
        self.root.geometry(f"{self.dimension_x1}x{self.dimension_y1}+{x}+{y}")
        self.root.title("ITESM ETI Prototype")

        self.root.resizable(False, False)
        self.root.configure(bg=self.used_color)
        self.canvas = self.create_canvas(self.dimension_x1, self.dimension_y1)
        self.serial_widgets()
        self.create_combo_widget()
        self.apply_combox_changes()
        self.leg_animation = LegAnimation(self.canvas, self.frames_path)
        self.init_widgets()

    def create_canvas(self, dimension_x, dimension_y):
        canvas = tk.Canvas(
            self.root,
            bg=self.used_color,
            height=int(dimension_y),
            width=int(dimension_x),
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )
        canvas.place(x=0, y=0)
        return canvas

    def serial_widgets(self):

        self.status_label = tk.Label(self.canvas, text="Sin conexión", fg=self.disconnected_color, font=("Inter", 12),
                                     bg=self.used_color)
        self.status_label.place(x=75.0, y=620.0)

        self.connect_button_image = PhotoImage(file=relative_to_assets("CONNECT_BTN0.png"))
        self.disconnect_button_image = PhotoImage(file=relative_to_assets("DISCONNECT_BTN0.png"))

        self.toggle_connection_button = Button(  # Botón de conexión
            self.canvas,
            image=self.connect_button_image,
            command=self.toggle_connection,
            relief="flat",
            bg=self.used_color
        )
        self.toggle_connection_button.place(x=70, y=550)

    """Cambia el estado de la conexión"""

    def toggle_connection(self):
        if self.is_connected:
            self.disconnect_arduino()
        else:
            self.connect_to_arduino()

    """Selección automáticadel  puerto conectado a Arduino"""

    def find_arduino_port(self):
        ports = serial.tools.list_ports.comports()
        for port in ports:
            if "Arduino" in port.description or "CH340" in port.description or "USB Serial" in port.description:
                return port.device
        return None

    """Lectura de datos seriales y cambbio de posición de animación"""

    def read_serial_port(self):
        try:
            while not self.stop_threads:
                if self.ser and self.ser.is_open:
                    data = self.ser.readline().decode().strip()
                    if data:
                        values = data.split(",")
                        if len(values) >= 2:
                            try:
                                rad = abs(float(values[0]))
                                self.position = (rad * 180) / math.pi
                                self.torque = abs(float(values[1]))
                                self.grados = self.position
                                if self.leg_animation:
                                    self.leg_animation.update_frame(self.position, self.torque)
                            except ValueError:
                                print("Error: Invalid data format. Skipping this line.")
                        else:
                            print("Error: Insufficient data. Skipping this line.")
                    else:
                        print("Warning: No data received.")
                else:
                    break
        except Exception as e:
            print("Error reading the serial port:", e)
            self.stop_threads = True

    """Conexión a arduino e inicialización de threads de lectura"""

    def connect_to_arduino(self):
        arduino_port = self.find_arduino_port()
        if arduino_port:
            if self.ser is None:
                try:
                    self.arduino_lock = threading.Lock()
                    self.ser = serial.Serial(arduino_port, 115200, timeout=2)
                    self.status_label.config(text=f"Conectado a: {arduino_port}", fg=self.connected_color,
                                             font=("Inter", 12))
                    self.toggle_connection_button.config(image=self.disconnect_button_image)
                    self.combobox.config(state="normal")
                    self.apply_button_widget.config(state="normal")
                    self.is_connected = True
                    self.stop_threads = False

                except Exception as e:
                    messagebox.showerror("Error", f"Failed to connect: {e}")
                    if self.ser:
                        self.ser.close()
                        self.ser = None
                if self.ser is not None:
                    reading_thread = threading.Thread(target=self.read_serial_port, args=self.ser)
                    reading_thread.start()
        else:
            messagebox.showwarning("Not Found", "Arduino not found. Please check the connection.")

    """Desconexión de arduino y detención de threads de lectura"""

    def disconnect_arduino(self):
        self.turn_off_motor()
        if self.ser and self.ser.is_open:
            self.stop_threads = True
            self.ser.close()
            self.ser = None
            self.status_label.config(text="Sin conexión", fg=self.disconnected_color, font=("Inter", 12))
            self.toggle_connection_button.config(image=self.connect_button_image)
            self.combobox.config(state="disabled")
            self.apply_button_widget.config(state="disabled")
            self.is_connected = False
            self.stop_threads = False

    """Inicialización de temporizador"""

    def start_timer(self):
        guide = (self.patient_data.get("Actividad", "No Registrado")).lower()
        key_id = guide[0:3]
        if not self.running:
            self.time_left = int(read_input_from_json(f"{key_id}_time"))  # Asigna la duración del temporizador
            self.running = True
            self.message_shown = False
            self.update_timer()

    """Detiene el temporizador"""

    def stop_timer(self):
        if self.running:
            self.time_left = 0  # Termina el temporizador
            self.running = False

    """Actualiza el temporizador y determina el resultado del estudio"""

    def update_timer(self):
        if not self.running:
            return

        if self.time_left > 0:
            minutes, seconds = divmod(self.time_left, 60)
            self.label_timer.config(text=f"{minutes:02}:{seconds:02}")
            self.time_left -= 1

            if self.position >= self.max_angle and not self.message_shown:
                # Posición objetivo alcanzada
                self.running = False
                self.achieved_test("#06D7A0")
                self.label_timer.config(text="00:00")
                messagebox.showinfo("Timer", "Ha alcanzado la posición deseada!")
                self.message_shown = True

            if self.running:
                self.root.after(1000, self.update_timer)

        elif self.time_left == 0 and self.position <= self.max_angle and not self.message_shown:
            # Posición objetivo no alcanzada
            self.running = False
            self.failed_test("#F04770")
            self.label_timer.config(text="00:00")
            messagebox.showinfo("Timer", "No ha alcanzado la posición deseada!")
            self.message_shown = True

    """ Lógica ejecutada con el botón iniciar. """

    def toggle_boton(self):
        if self.boton_toggle["text"] == "Iniciar":
            # Se cambia la imagen del botón y se inicializa la animación del indicador y el envío de datos seriales
            messagebox.showinfo("Estudio", "Se ha comenzado a aplicar fuerza!")
            self.animation_on_write_serial()
            self.boton_toggle.config(text="Detener", image=self.imagen_detener, command=self.toggle_boton)
        else:
            # Se cambia la imagen del botón y se detiene la animación, el timer, y el thread de envío de datos seriales
            self.animation_off_write_serial()
            self.stop_timer()
            self.label_timer.config(text="00:00")
            self.boton_toggle.config(text="Iniciar", image=self.imagen_iniciar, command=self.toggle_boton)

            # Detiene el ciclo de envío de datos
            self.send_value_running = False
            if self.send_value_thread is not None:
                self.send_value_thread.join()  # Espera a que el thread de envío de datos termine
                self.send_value_thread = None

            # Reinicia los colores del indicador de niveles obtenidos/no obtenidos
            if self.level.startswith("Nivel "):
                level_num = int(self.level.split(" ")[1])
                self.squares[level_num - 1].config(bg="#FFFFFF")
                self.achieved_levels[level_num - 1] = False

    """ Envío de datos seriales """

    def send_value(self):
        cadena = str(self.combobox.get())
        nivel = int(cadena.split()[1])
        guide = (self.patient_data.get("Actividad", "No Registrado")).lower()
        key_id = guide[0:3]

        if nivel <= 3:
            cadena = f"{key_id}_nv{nivel}"
            print(cadena)
            final_value = int(read_input_from_json(cadena))
            divided_value = final_value / 4
            cumulative_value = 0

            while self.send_value_running and cumulative_value < final_value:
                cumulative_value += divided_value
                if self.ser and self.ser.is_open:  # Ensure the serial port is open
                    self.arduino_lock.acquire()  # Acquire the lock for thread-safe access
                    try:
                        self.ser.write((str(cumulative_value) + "\n").encode('ascii'))  # Send the cumulative value
                        print(f"Sent: {cumulative_value}")  # Debug print
                    except Exception as e:
                        print(f"Error writing to serial port: {e}")
                    finally:
                        self.arduino_lock.release()  # Release the lock
                else:
                    print("Serial port is not open.")
                    break

                time.sleep(2)  # Wait for 2 seconds before the next iteration

            # When max torque is reached
            if cumulative_value >= final_value:
                self.max_torque_reached = True
                messagebox.showinfo("Alerta", "Se ha alcanzado el torque máximo.")
                self.start_timer()

        elif (int(nivel) >= 4) and self.user_input.get().isdigit():
            cadena = str(self.user_input.get())
            nivel = int(cadena)
            divided_value = nivel / 4
            cumulative_value = 0
            while self.send_value_running and cumulative_value < nivel:
                cumulative_value += divided_value
                if self.ser and self.ser.is_open:  # Ensure the serial port is open
                    self.arduino_lock.acquire()  # Acquire the lock for thread-safe access
                    try:
                        self.ser.write((str(cumulative_value) + "\n").encode('ascii'))  # Send the cumulative value
                        print(f"Sent: {cumulative_value}")  # Debug print
                    except Exception as e:
                        print(f"Error writing to serial port: {e}")
                    finally:
                        self.arduino_lock.release()  # Release the lock
                else:
                    print("Serial port is not open.")
                    break

                time.sleep(2)  # Wait for 2 seconds before the next iteration

            # When max torque is reached
            if cumulative_value >= nivel:
                self.max_torque_reached = True
                messagebox.showinfo("Alerta", "Se ha alcanzado el torque máximo.")
                self.start_timer()

    """ Inicia la animación de barra, enciende el motor e inicia el envío de datos seriales """

    def animation_on_write_serial(self):
        self.combobox.config(state="disabled")
        self.start_animation()
        time.sleep(0.10)
        self.turn_on_motor()
        time.sleep(0.10)

        # Start the send_value function in a separate thread
        self.send_value_running = True
        self.max_torque_reached = False
        self.send_value_thread = threading.Thread(target=self.send_value)
        self.send_value_thread.start()

    def animation_off_write_serial(self):
        self.combobox.config(state="readonly")
        self.stop_animation()
        time.sleep(0.05)
        self.turn_off_motor()
        time.sleep(0.05)

    def turn_on_motor(self):
        if self.ser is not None and self.ser.is_open:
            cadena = "998\n"
            self.arduino_lock.acquire()
            try:
                self.ser.write(cadena.encode('ascii'))
                print("Motor turned on.")
            except Exception as e:
                print(f"Error turning on motor: {e}")
            finally:
                self.arduino_lock.release()

    def turn_off_motor(self):
        if self.ser is not None and self.ser.is_open:
            cadena = "999\n"
            self.arduino_lock.acquire()
            try:
                self.ser.write(cadena.encode('ascii'))
                print("Motor turned off.")
            except Exception as e:
                print(f"Error turning off motor: {e}")
            finally:
                self.arduino_lock.release()

    def origin_motor(self):
        if self.ser is not None:
            cadena = "997\n"
            self.arduino_lock.acquire()
            time.sleep(0.05)
            self.ser.write(cadena.encode('ascii'))
            time.sleep(0.05)
            self.arduino_lock.release()

    def validate_input(self, new_value):
        return new_value.isdigit() or new_value == ""

    def update_state(self, event=None):
        if self.combobox.get() in ["Nivel 4", "Nivel 5"]:
            self.user_input.config(state="normal")
        else:
            self.user_input.config(state="disabled")
            self.user_input.delete(0, tk.END)

    def check_last_lvl(self, actual_lvl):
        if actual_lvl > 1 and not self.achieved_levels[actual_lvl - 1]:
            answer = messagebox.askquestion("Confirmación",
                                            f"¿Pasó exitosamente los niveles 1 a {actual_lvl - 1}?")
            if answer == "yes":
                for i in range(actual_lvl - 1):
                    self.squares[i - 5].config(bg="#06D7A0")
                    self.achieved_levels[i] = True
                return True
            else:
                messagebox.showwarning("Aviso",
                                       "Por favor, seleccione el nivel faltante para hacer los test de forma correcta.")
                self.combobox.set("Seleccionar el nivel")
                return False
        return True

    def create_combo_widget(self):
        # Title
        self.canvas.create_text(350, 10, anchor="nw", text="Estudio automático", fill=self.text_color,
                                font=("Inter", 30))
        levels = ["Nivel 1", "Nivel 2", "Nivel 3", "Nivel 4", "Nivel 5"]
        self.combobox = ttk.Combobox(self.canvas, values=levels, font=("Inter", 14), width=20)
        self.combobox.set("Elija el nivel de fuerza")
        self.combobox.config(state="disabled")  # Make the combobox readonly
        self.combobox.place(x=360, y=555)

        self.cmd_entry = self.canvas.register(self.validate_input)
        self.user_input = tk.Entry(self.canvas, font=("Inter", 16), width=10, validate="key",
                                   validatecommand=(self.cmd_entry, "%P"), bg="white")
        self.user_input.place(x=420, y=600)
        self.user_input.config(state="disabled")

        self.apply_image = PhotoImage(file=relative_to_assets("APPLY_BTN0.png"))
        self.apply_button_widget = tk.Button(self.canvas, image=self.apply_image, command=self.apply_combox_changes,
                                             relief="flat", borderwidth=0, bg=self.used_color)
        self.apply_button_widget.place(x=380, y=640)
        self.apply_button_widget.config(state="disabled")

        self.mensaje_label1 = tk.Label(self.canvas, text="", font=("Inter", 12), bg=self.used_color)
        self.mensaje_label1.place(x=590, y=640)

        self.strengthT_label = tk.Label(self.canvas, text="F =", font=("Inter", 18), bg=self.used_color, fg="#000000")
        self.strengthT_label.place(x=355, y=600)
        self.strengthKG_label = tk.Label(self.canvas, text="N/m", font=("Inter", 18), bg=self.used_color, fg="#000000")
        self.strengthKG_label.place(x=555, y=600)

        # self.nivelesF_label = self.canvas.create_text(420, 530, text="Niveles de fuerza", font=("Inter", 13), fill=self.text_color)

        self.indicator1_bg_image = PhotoImage(file=relative_to_assets("INDICATOR1_BG0.png"))
        self.canvas.create_image(709, 135, image=self.indicator1_bg_image, anchor="nw")

        self.origin_bg_image = PhotoImage(file=relative_to_assets("SET_ORIGIN_BTN_BG.png"))
        self.origin_btn = tk.Button(self.canvas, image=self.origin_bg_image, command=self.origin_motor,
                                    relief="flat", borderwidth=0, bg=self.used_color)
        self.origin_btn.place(x=735.0, y=160.0)

        self.max_angle_text = self.canvas.create_text(810, 270, text="Ángulo máx.:  0.0 °", font=("Inter", 12),
                                                      fill=self.text_color)

        self.save_angle_bg = PhotoImage(file=relative_to_assets("SAVE_ANGLE_BTN0_BG.png"))

        self.max_angle_btn = tk.Button(self.canvas, image=self.save_angle_bg, command=self.save_max_angle,
                                       relief="flat", borderwidth=0, bg=self.used_color)
        self.max_angle_btn.place(x=735.0, y=284.0)
        self.max_angle_btn.config(state="disabled")

        self.label_timer = tk.Label(root, text="00:00", font=("Inter", 12), bg=self.used_color)
        self.label_timer.place(x=740.0, y=385.0)

        self.leg_bg_image = PhotoImage(file=relative_to_assets("LEG_BG.png"))
        self.canvas.create_image(40, 134, image=self.leg_bg_image, anchor="nw")

        self.inidcator_bg_image = PhotoImage(file=relative_to_assets("INDICATOR_BG.png"))
        self.canvas.create_image(40, 445, image=self.inidcator_bg_image, anchor="nw")

        self.inidcator_bg_image1 = PhotoImage(file=relative_to_assets("INDICATOR_BG.png"))
        self.canvas.create_image(300, 445, image=self.inidcator_bg_image1, anchor="nw")

        self.grados_text = self.canvas.create_text(100, 470, text="Grados:", font=("Inter", 13), fill=self.text_color)
        self.torque_text = self.canvas.create_text(360, 470, text="Torque:", font=("Inter", 13), fill=self.text_color)

        self.gauge_bg_image = PhotoImage(file=relative_to_assets("GAUGE_BG0.png"))
        self.canvas.create_image(562, 135, image=self.gauge_bg_image, anchor="nw")

        for i in range(5):
            self.square_widget = tk.Label(self.canvas, text=str(i + 1), font=("Inter", 10), width=12, height=4, bd=0,
                                          highlightbackground="#D3D4D9",
                                          highlightthickness=1, bg="white")
            self.square_widget.place(x=583, y=398 - (i * 60))
            self.squares.append(self.square_widget)

        self.titleC_label = self.canvas.create_text(570, 120, text="Niveles", font=("Inter", 13), fill=self.text_color
                                                    , width=150, anchor="w")

        self.imagen_iniciar = PhotoImage(file=relative_to_assets("START_BTN0.png"))
        self.imagen_detener = PhotoImage(file=relative_to_assets("STOP_BTN0.png"))

        self.save_image = PhotoImage(file=relative_to_assets("SAVE_BTN0.png"))
        self.boton_save = tk.Button(
            self.canvas,
            image=self.save_image,
            state="normal",
            relief="flat",
            borderwidth=0,
            bg=self.used_color,
            command=self.save_boton)
        self.boton_save.place(x=720, y=550)
        self.boton_save.config(state="disabled")

        self.boton_toggle = tk.Button(
            self.canvas,
            text="Iniciar",
            font=("Inter", 16),
            command=self.toggle_boton,
            state="normal",
            relief="flat",
            borderwidth=0,
            bg=self.used_color,
            image=self.imagen_iniciar)
        self.boton_toggle.place(x=734, y=410)
        self.boton_toggle.config(state="disabled")

        self.combobox.bind("<<ComboboxSelected>>", self.update_state)

        self.achieved_levels = [False] * 5

    def save_max_angle(self):
        if self.position is not None:
            self.canvas.itemconfig(self.max_angle_text, text=f"Ángulo máx.:  {self.position:.1f} °")
            self.max_angle = self.position
        else:
            messagebox.showinfo("Error", "No se detecta posición, verifique el sistema")

    def achieved_test(self, color):
        nivel_actual = self.combobox.get()
        if nivel_actual in self.datos:
            self.datos[nivel_actual].append({
                "Grados": self.grados,
                "Torque": self.torque,
                "Llegó": "Sí" if color == "#06D7A0" else "No"  # Dependiendo del color, sabes si llegó o no
            })
        self.highlight(color)
        self.combobox.config(state="readonly")
        self.turn_off_motor()

    def failed_test(self, color):
        nivel_actual = self.combobox.get()
        if nivel_actual in self.datos:
            self.datos[nivel_actual].append({
                "Grados": self.grados,
                "Torque": self.torque,
                "Llegó": "Sí" if color == "#06D7A0" else "No"  # Dependiendo del color, sabes si llegó o no
            })
        self.highlight(color)
        self.combobox.config(state="readonly")
        self.turn_off_motor()

    def apply_combox_changes(self):
        self.level = self.combobox.get()
        self.value = self.user_input.get()
        guide = (self.patient_data.get("Actividad", "No Registrado")).lower()
        key_id = guide[0:3]
        lvl4_max = int(read_input_from_json(f"{key_id}_nv4"))
        lvl5_max = int(read_input_from_json(f"{key_id}_nv5"))
        if self.level.startswith("Nivel "):
            self.level_num = int(self.level.split(" ")[1])

            if self.level_num == 4 and self.value.isdigit() and int(self.value) > lvl4_max:
                self.mensaje_label1.config(text=f"ERROR: Máx. {lvl4_max} N/m", fg="#00072d", bg="#FFFFFF")
                return
            elif self.level_num == 5 and self.value.isdigit() and int(self.value) > lvl5_max:
                self.mensaje_label1.config(text=f"ERROR: Máx. {lvl5_max} N/m", fg="#00072d", bg="#FFFFFF")
                return
            elif self.level_num in (4, 5) and ((not self.value.strip() or not self.value.isdigit())
                                               or int(self.value) == 0):
                self.mensaje_label1.config(text="ERROR: valor inválido", fg="#00072d", bg="#FFFFFF")
                return
            if not self.check_last_lvl(self.level_num):
                return

            self.mensaje_label1.config(text="Cambios aplicados", fg="#00072D")
            self.user_input.config(state="disabled")
            self.boton_toggle.config(state="normal")
            self.combobox.config(state="disabled")
            self.max_angle_btn.config(state="normal")

        self.mensaje_label1.after(5000, lambda: self.mensaje_label1.config(text=""))

    def auto_size_columns(ws):
        """
        Ajusta automáticamente el ancho de las columnas en una hoja de Excel.
        """
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Ajustar el ancho con un pequeño margen
            ws.column_dimensions[column].width = adjusted_width

    def save_boton(self):
        try:
            # Crear una lista para almacenar todos los datos
            datos_finales = []

            # Recorrer el diccionario y agregar los datos a la lista
            for nivel, registros in self.datos.items():
                for registro in registros:
                    datos_finales.append({
                        "Nivel": nivel,
                        "Grados": registro["Grados"],
                        "Torque": registro["Torque"],
                        "Llegó": registro["Llegó"]
                    })

            # Obtener la ruta de salida y el número de expediente desde patient_data
            expediente = self.patient_data.get("Expediente", "Sin expediente")
            output_folder = self.patient_data.get("output_folder", Path.cwd())

            if expediente and output_folder:
                # Crear el nombre del archivo con el número de expediente
                output_path = output_folder / f"{expediente}.xlsx"

                # Verificar si el archivo ya existe
                if output_path.exists():
                    # Abrir el archivo existente
                    wb = openpyxl.load_workbook(output_path)
                    # Verificar si ya existe una hoja llamada "Modo Automático"
                    if "Modo Automático" in wb.sheetnames:
                        # Si existe, eliminarla para evitar duplicados
                        wb.remove(wb["Modo Automático"])
                    # Crear una nueva hoja para el modo automático
                    ws_auto = wb.create_sheet("Modo Automático")
                else:
                    # Si el archivo no existe, crear uno nuevo
                    wb = openpyxl.Workbook()
                    # Crear la hoja para el modo automático
                    ws_auto = wb.active
                    ws_auto.title = "Modo Automático"

                # Escribir la tabla de información personal
                self.write_patient_info(ws_auto)

                # Escribir la tabla de resumen de pruebas de fuerza
                self.write_test_summary(ws_auto, datos_finales)

                # Crear la gráfica de barras
                chart = BarChart()
                chart.title = "Grados por Nivel"
                chart.x_axis.title = "Nivel"  # Eje X: Niveles
                chart.y_axis.title = "Grados"  # Eje Y: Grados

                # Referencias para los datos de la gráfica
                data = Reference(ws_auto, min_col=2, min_row=8,
                                 max_row=8 + len(datos_finales))  # Columna B (Grados)
                categories = Reference(ws_auto, min_col=1, min_row=8,
                                       max_row=8 + len(datos_finales))  # Columna A (Nivel)

                # Agregar los datos a la gráfica
                chart.add_data(data, titles_from_data=False)  # titles_from_data=False para evitar usar B8 como título
                chart.set_categories(categories)

                # Configurar las etiquetas de datos (valores arriba de las barras)
                chart.dLbls = None  # Deshabilitar las etiquetas de datos

                # Agregar la gráfica a la hoja
                ws_auto.add_chart(chart, "F7")  # Colocar la gráfica al lado de la tabla

                # Guardar el archivo Excel
                wb.save(output_path)

                # Mostrar mensaje de éxito
                messagebox.showinfo("Test Finalizado",
                                    f"El registro y test ha concluido. Su archivo ha sido guardado como {output_path}.")
            else:
                messagebox.showwarning("Error",
                                       "Por favor, ingrese un número de expediente válido y seleccione una carpeta de salida.")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al guardar el archivo: {e}")

        # Deshabilitar el botón de guardar y resetear la interfaz
        self.boton_save.config(state="disabled")
        self.user_input.config(state="disabled")
        self.squares[4].config(bg="#FFFFFF")
        self.squares[3].config(bg="#FFFFFF")
        self.squares[2].config(bg="#FFFFFF")
        self.squares[1].config(bg="#FFFFFF")
        self.squares[0].config(bg="#FFFFFF")
        self.achieved_levels[:] = [False] * len(self.achieved_levels)
        self.datos = {
            "Nivel 1": [],
            "Nivel 2": [],
            "Nivel 3": [],
            "Nivel 4": [],
            "Nivel 5": []
        }

    def write_patient_info(self, ws):
        """Escribe la información del paciente en la hoja de Excel."""
        ws['A1'] = "Información Personal"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:F1')  # Combinar celdas para el header (ahora son 5 columnas)

        # Subheaders de la tabla de información personal
        ws['A2'] = "Nombre"
        ws['B2'] = "Edad"
        ws['C2'] = "Sexo"
        ws['D2'] = "Actividad Física"
        ws['E2'] = "Fecha de prueba"
        ws['F2'] = "Ángulo máximo de prueba"
        for cell in ws['A2:F2'][0]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos de la tabla de información personal
        ws['A3'] = self.patient_data.get("Nombre", "No registrado")
        ws['B3'] = self.patient_data.get("Edad", "No registrado")
        ws['C3'] = self.patient_data.get("Sexo", "No registrado")
        ws['D3'] = self.patient_data.get("Actividad", "No registrado")
        ws['E3'] = self.patient_data.get("Fecha de prueba",
                                         "No registrado")  # Puedes agregar un campo para la fecha si es necesario
        ws['F3'] = self.max_angle
        for cell in ws['A3:F3'][0]:
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def write_test_summary(self, ws, datos_finales):
        """Escribe el resumen de las pruebas de fuerza en la hoja de Excel."""
        # Espacio entre las tablas
        ws['A5'] = ""  # Espacio vacío

        # Escribir la tabla de resumen de pruebas de fuerza
        ws['A6'] = "Resumen de Pruebas de Fuerza"
        ws['A6'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A6'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws['A6'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A6:D6')  # Combinar celdas para el header

        # Subheaders de la tabla de resumen de pruebas de fuerza
        ws['A7'] = "Nivel"
        ws['B7'] = "Grados"
        ws['C7'] = "Torque"
        ws['D7'] = "Llegó"
        for cell in ws['A7:D7'][0]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos de la tabla de resumen de pruebas de fuerza
        row_index = 8
        for dato in datos_finales:
            ws[f'A{row_index}'] = dato["Nivel"]
            ws[f'B{row_index}'] = dato["Grados"]
            ws[f'C{row_index}'] = dato["Torque"]
            ws[f'D{row_index}'] = dato["Llegó"]
            for cell in ws[f'A{row_index}:D{row_index}'][0]:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_index += 1

    def start_animation(self):
        self.level1 = self.combobox.get()
        if self.level1.startswith("Nivel "):
            self.level1_num = int(self.level1.split(" ")[1])
            if 1 <= self.level1_num <= 5:
                self.active_animation = True
                self.blink_state = True
                self.blinking(self.squares[self.level1_num - 1])
                self.combobox.config(state="disabled")
                self.user_input.config(state="disabled")
                self.boton_save.config(state="disabled")
                self.boton_toggle.config(text="Detener", image=self.imagen_detener)

    def blinking(self, square):
        if self.active_animation:
            color = "#60AFFF" if self.blink_state else "white"
            square.config(bg=color)
            self.blink_state = not self.blink_state
            square.after(500, lambda: self.blinking(square))

    def stop_animation(self):
        self.level = self.combobox.get()
        self.active_animation = False
        self.combobox.config(state="normal")
        self.boton_toggle.config(text="Iniciar", image=self.imagen_iniciar, state="disabled")
        self.boton_save.config(state="normal")

        if self.level.startswith("Nivel "):
            self.nivel_num = int(self.level.split(" ")[1])
            if self.nivel_num > 3:
                self.user_input.config(state="normal")

    def highlight(self, color):
        self.level = self.combobox.get()

        if self.level.startswith("Nivel "):
            self.level_num = int(self.level.split(" ")[1])
            if 1 <= self.level_num <= 3:
                self.stop_animation()
                self.squares[self.level_num - 1].config(bg=color)
                self.achieved_levels[self.level_num - 1] = True
                self.user_input.config(state="disabled")
            else:
                self.stop_animation()
                self.squares[self.level_num - 1].config(bg=color)
                self.achieved_levels[self.level_num - 1] = True
                self.user_input.config(state="normal")

    def init_widgets(self):
        # Mostrar información del paciente
        nombre_paciente = self.patient_data.get("Nombre", "No registrado")
        self.nombre_title = self.canvas.create_text(50, 120, text=f"Pierna de: {nombre_paciente}",
                                                    font=("Inter", 13), fill=self.text_color, width=400, anchor="w")

        self.return_image = PhotoImage(file=relative_to_assets("GO_BACK_BTN1.png"))
        self.return_btn = tk.Button(self.canvas, state="normal", relief="flat",
                                    borderwidth=0,
                                    bg=self.used_color,
                                    image=self.return_image,
                                    command=lambda: self.controller.switch_frame(AppInterface1))
        self.return_btn.place(x=810, y=645)

    def on_closing(self):
        self.turn_off_motor()
        time.sleep(0.05)
        self.disconnect_arduino()
        self.root.destroy()


"""
############################################## QUINTA VENTANA ##########################################################
Ventana que realiza el estudio MANUAL

########################################################################################################################
"""


class AppInterface3(AppBase):
    def __init__(self, root, controller, patient_data, config_data):
        super().__init__(root, controller, patient_data, config_data)
        self.updateMeterLine = None
        self.MeterWidget = None
        self.messagebox = None
        self.stop_button_widget = None
        self.start_button_widget = None
        self.send_data = None
        self.level = None
        self.value = None
        self.text = None
        self.combobox = None
        self.entry_bg_1 = None
        self.entry_1 = None
        self.connect_button_image = None
        self.leg_animation = None
        self.ser = None
        self.type_input = None
        self.stop_threads = False
        self.is_connected = False
        self.active_animation = True
        self.blink_state = True
        self.grados = 0
        self.torque = 0
        self.nivel_actual = None
        self.datos = {
            "Nivel 1": [],
            "Nivel 2": [],
            "Nivel 3": [],
            "Nivel 4": [],
            "Nivel 5": []
        }
        self.frames_path = Path(__file__).resolve().parent / "light_video"
        self.squares = []
        self.root = root
        self.dimension_x1 = "1000"
        self.dimension_y1 = "720"

        self.send_value_running = False
        self.send_value_thread = None
        self.max_torque_reached = False

        self.screen_width = self.root.winfo_screenwidth()  # Obtiene el ancho de la pantalla
        self.screen_height = self.root.winfo_screenheight()  # Obtiene el alto de la pantalla
        x = (self.screen_width // 2) - (int(self.dimension_x1) // 2)  # Calcula la posición X
        y = (self.screen_height // 2) - (int(self.dimension_y1) // 2)  # Calcula la posición Y
        self.root.geometry(f"{self.dimension_x1}x{self.dimension_y1}+{x}+{y}")

        self.root.resizable(False, False)
        self.root.configure(bg=self.used_color)
        self.canvas = self.create_canvas(self.dimension_x1, self.dimension_y1)
        self.serial_widgets()
        self.create_combo_widget()
        self.apply_combox_changes()
        self.leg_animation = LegAnimation(self.canvas, self.frames_path)
        self.init_widgets()
        self.root.title("ITESM ETI Prototype")

    def create_canvas(self, dimension_x, dimension_y):
        canvas = tk.Canvas(
            self.root,
            bg=self.used_color,
            height=int(dimension_y),
            width=int(dimension_x),
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )
        canvas.place(x=0, y=0)
        return canvas

    def serial_widgets(self):

        self.status_label = tk.Label(self.canvas, text="Sin conexión", fg=self.disconnected_color, font=("Inter", 12),
                                     bg=self.used_color)
        self.status_label.place(x=75.0, y=620.0)

        self.connect_button_image = PhotoImage(file=relative_to_assets("CONNECT_BTN0.png"))
        self.disconnect_button_image = PhotoImage(file=relative_to_assets("DISCONNECT_BTN0.png"))

        self.toggle_connection_button = Button(
            self.canvas,
            image=self.connect_button_image,
            command=self.toggle_connection,
            relief="flat",
            bg=self.used_color
        )
        self.toggle_connection_button.place(x=70, y=550)

    def toggle_connection(self):
        if self.is_connected:
            self.disconnect_arduino()
        else:
            self.connect_to_arduino()

    def find_arduino_port(self):
        ports = serial.tools.list_ports.comports()
        for port in ports:
            if "Arduino" in port.description or "CH340" in port.description or "USB Serial" in port.description:
                return port.device
        return None

    def read_serial_port(self):
        try:
            while not self.stop_threads:
                if self.ser and self.ser.is_open:
                    data = self.ser.readline().decode().strip()
                    if data:
                        values = data.split(",")
                        if len(values) >= 2:
                            try:
                                rad = abs(float(values[0]))
                                self.position = (rad * 180) / math.pi
                                self.torque = abs(float(values[1]))
                                self.grados = self.position
                                if self.leg_animation:
                                    self.leg_animation.update_frame(self.position, self.torque)
                            except ValueError:
                                print("Error: Invalid data format. Skipping this line.")
                        else:
                            print("Error: Insufficient data. Skipping this line.")
                    else:
                        print("Warning: No data received.")
                else:
                    break
        except Exception as e:
            print("Error reading the serial port:", e)
            self.stop_threads = True

    def send_serial_port(self):
        try:
            while not self.stop_threads is True:
                time.sleep(0.05)
        except Exception as e:
            print("Error sending data", e)
            self.stop_threads = True

    def connect_to_arduino(self):
        arduino_port = self.find_arduino_port()
        if arduino_port:
            if self.ser is None:
                try:
                    self.arduino_lock = threading.Lock()
                    self.ser = serial.Serial(arduino_port, 115200, timeout=2)
                    self.status_label.config(text=f"Conectado a: {arduino_port}", fg=self.connected_color,
                                             font=("Inter", 12))
                    self.toggle_connection_button.config(image=self.disconnect_button_image)
                    self.combobox.config(state="normal")
                    self.apply_button_widget.config(state="normal")
                    self.is_connected = True
                    self.stop_threads = False

                except Exception as e:
                    messagebox.showerror("Error", f"Failed to connect: {e}")
                    if self.ser:
                        self.ser.close()
                        self.ser = None
                if self.ser is not None:
                    # Starting the reading thread
                    reading_thread = threading.Thread(target=self.read_serial_port, args=self.ser)
                    reading_thread.start()

        else:
            messagebox.showwarning("Not Found", "Arduino not found. Please check the connection.")

    def disconnect_arduino(self):
        self.turn_off_motor()
        if self.ser and self.ser.is_open:
            self.stop_threads = True
            self.ser.close()
            self.ser = None
            self.status_label.config(text="Sin conexión", fg=self.disconnected_color, font=("Inter", 12))
            self.toggle_connection_button.config(image=self.connect_button_image)
            self.combobox.config(state="disabled")
            self.apply_button_widget.config(state="disabled")
            self.is_connected = False
            self.stop_threads = False

    def toggle_boton(self):
        if self.boton_toggle["text"] == "Iniciar":
            messagebox.showinfo("Estudio", "Se ha comenzado a aplicar fuerza!")
            self.animation_on_write_serial()
            self.boton_toggle.config(text="Detener", image=self.imagen_detener, command=self.toggle_boton)
        else:
            self.animation_off_write_serial()
            self.boton_toggle.config(text="Iniciar", image=self.imagen_iniciar, command=self.toggle_boton)

            # Stop the send_value loop
            self.send_value_running = False
            if self.send_value_thread is not None:
                self.send_value_thread.join()  # Wait for the thread to finish
                self.send_value_thread = None

            # Reset the squares and achieved levels if necessary
            if self.level.startswith("Nivel "):
                level_num = int(self.level.split(" ")[1])
                self.squares[level_num - 1].config(bg="#FFFFFF")
                self.achieved_levels[level_num - 1] = False

    def send_value(self):
        cadena = str(self.combobox.get())
        nivel = int(cadena.split()[1])
        guide = (self.patient_data.get("Actividad", "No Registrado")).lower()
        key_id = guide[0:3]

        if nivel <= 3:
            cadena = f"{key_id}_nv{nivel}"
            final_value = int(read_input_from_json(cadena))
            divided_value = final_value / 4
            cumulative_value = 0

            while self.send_value_running and cumulative_value < final_value:
                cumulative_value += divided_value
                if self.ser and self.ser.is_open:  # Ensure the serial port is open
                    self.arduino_lock.acquire()  # Acquire the lock for thread-safe access
                    try:
                        self.ser.write((str(cumulative_value) + "\n").encode('ascii'))  # Send the cumulative value
                        print(f"Sent: {cumulative_value}")  # Debug print
                    except Exception as e:
                        print(f"Error writing to serial port: {e}")
                    finally:
                        self.arduino_lock.release()  # Release the lock
                else:
                    print("Serial port is not open.")
                    break

                time.sleep(2)  # Wait for 2 seconds before the next iteration

            # When max torque is reached
            if cumulative_value >= final_value:
                self.max_torque_reached = True
                messagebox.showinfo("Alerta", "Se ha alcanzado el torque máximo.")

        elif (int(nivel) >= 4) and self.user_input.get().isdigit():
            cadena = str(self.user_input.get())
            nivel = int(cadena)
            divided_value = nivel / 4
            cumulative_value = 0
            while self.send_value_running and cumulative_value < nivel:
                cumulative_value += divided_value
                if self.ser and self.ser.is_open:  # Ensure the serial port is open
                    self.arduino_lock.acquire()  # Acquire the lock for thread-safe access
                    try:
                        self.ser.write((str(cumulative_value) + "\n").encode('ascii'))  # Send the cumulative value
                        print(f"Sent: {cumulative_value}")  # Debug print
                    except Exception as e:
                        print(f"Error writing to serial port: {e}")
                    finally:
                        self.arduino_lock.release()  # Release the lock
                else:
                    print("Serial port is not open.")
                    break

                time.sleep(2)  # Wait for 2 seconds before the next iteration

            # When max torque is reached
            if cumulative_value >= nivel:
                self.max_torque_reached = True
                messagebox.showinfo("Alerta", "Se ha alcanzado el torque máximo.")

    def animation_on_write_serial(self):
        self.combobox.config(state="disabled")
        self.start_animation()
        time.sleep(0.10)
        self.turn_on_motor()
        time.sleep(0.10)

        # Start the send_value function in a separate thread
        self.send_value_running = True
        self.max_torque_reached = False
        self.send_value_thread = threading.Thread(target=self.send_value)
        self.send_value_thread.start()

    def animation_off_write_serial(self):
        self.combobox.config(state="readonly")
        self.stop_animation()
        time.sleep(0.05)
        self.turn_off_motor()
        time.sleep(0.05)

    def turn_on_motor(self):
        if self.ser is not None:
            cadena = "998\n"
            self.arduino_lock.acquire()
            time.sleep(0.05)
            self.ser.write(cadena.encode('ascii'))
            time.sleep(0.05)
            self.arduino_lock.release()

    def turn_off_motor(self):
        if self.ser is not None:
            cadena = "999\n"
            self.arduino_lock.acquire()
            time.sleep(0.05)
            self.ser.write(cadena.encode('ascii'))
            time.sleep(0.05)
            self.arduino_lock.release()

    def origin_motor(self):
        if self.ser is not None:
            cadena = "997\n"
            self.arduino_lock.acquire()
            time.sleep(0.05)
            self.ser.write(cadena.encode('ascii'))
            time.sleep(0.05)
            self.arduino_lock.release()

    def validate_input(self, new_value):
        return new_value.isdigit() or new_value == ""

    def update_state(self, event=None):
        if self.combobox.get() in ["Nivel 4", "Nivel 5"]:
            self.user_input.config(state="normal")
        else:
            self.user_input.config(state="disabled")
            self.user_input.delete(0, tk.END)

    def check_last_lvl(self, actual_lvl):
        if actual_lvl > 1 and not self.achieved_levels[actual_lvl - 1]:
            answer = messagebox.askquestion("Confirmación",
                                            f"¿Pasó exitosamente los niveles 1 a {actual_lvl - 1}?")
            if answer == "yes":
                for i in range(actual_lvl - 1):
                    self.squares[i - 5].config(bg="#06D7A0")
                    self.achieved_levels[i] = True
                return True
            else:
                messagebox.showwarning("Aviso",
                                       "Por favor, seleccione el nivel faltante para hacer los test de forma correcta.")
                self.combobox.set("Seleccionar el nivel")
                return False
        return True

    def create_combo_widget(self):
        # Title
        self.canvas.create_text(380, 10, anchor="nw", text="Estudio Manual", fill=self.text_color, font=("Inter", 30))
        levels = ["Nivel 1", "Nivel 2", "Nivel 3", "Nivel 4", "Nivel 5"]
        self.combobox = ttk.Combobox(self.canvas, values=levels, font=("Inter", 14), width=20)
        self.combobox.set("Elija el nivel de fuerza")
        self.combobox.config(state="disabled")  # Make the combobox readonly
        self.combobox.place(x=360, y=555)

        self.cmd_entry = self.canvas.register(self.validate_input)
        self.user_input = tk.Entry(self.canvas, font=("Inter", 16), width=10, validate="key",
                                   validatecommand=(self.cmd_entry, "%P"), bg="white")
        self.user_input.place(x=420, y=600)
        self.user_input.config(state="disabled")

        self.apply_image = PhotoImage(file=relative_to_assets("APPLY_BTN0.png"))
        self.apply_button_widget = tk.Button(self.canvas, image=self.apply_image,
                                             command=self.apply_combox_changes, relief="flat", borderwidth=0,
                                             bg=self.used_color)
        self.apply_button_widget.place(x=380, y=640)
        self.apply_button_widget.config(state="disabled")

        self.mensaje_label1 = tk.Label(self.canvas, text="", font=("Inter", 12), bg=self.used_color)
        self.mensaje_label1.place(x=590, y=640)

        self.strengthT_label = tk.Label(self.canvas, text="F =", font=("Inter", 18), bg=self.used_color, fg="#404045")
        self.strengthT_label.place(x=355, y=600)
        self.strengthKG_label = tk.Label(self.canvas, text="N/m", font=("Inter", 18), bg=self.used_color, fg="#404045")
        self.strengthKG_label.place(x=555, y=600)

        self.indicator1_bg_image = PhotoImage(file=relative_to_assets("INDICATOR1_BG0.png"))
        self.canvas.create_image(709, 135, image=self.indicator1_bg_image, anchor="nw")

        self.achieved_bg_image = PhotoImage(file=relative_to_assets("ACHIEVED_BTN_BG.png"))
        self.achieved_button = Button(self.canvas, image=self.achieved_bg_image,
                                      command=lambda: self.achieved_test("#06D7A0"),
                                      relief="flat", bg=self.used_color, state="disabled")
        self.achieved_button.place(x=730.0, y=240.0)

        self.failed_bg_image = PhotoImage(file=relative_to_assets("FAILED_BTN_BG.png"))
        self.failed_button = Button(self.canvas, image=self.failed_bg_image,
                                    command=lambda: self.failed_test("#F04770"),
                                    relief="flat", bg=self.used_color, state="disabled")
        self.failed_button.place(x=730.0, y=325.0)

        self.origin_bg_image = PhotoImage(file=relative_to_assets("SET_ORIGIN_BTN_BG.png"))
        self.origin_btn = tk.Button(self.canvas, image=self.origin_bg_image, command=self.origin_motor,
                                    relief="flat", borderwidth=0, bg=self.used_color)
        self.origin_btn.place(x=735.0, y=160.0)

        self.leg_bg_image = PhotoImage(file=relative_to_assets("LEG_BG.png"))
        self.canvas.create_image(40, 134, image=self.leg_bg_image, anchor="nw")

        self.inidcator_bg_image = PhotoImage(file=relative_to_assets("INDICATOR_BG.png"))
        self.canvas.create_image(40, 445, image=self.inidcator_bg_image, anchor="nw")

        self.inidcator_bg_image1 = PhotoImage(file=relative_to_assets("INDICATOR_BG.png"))
        self.canvas.create_image(300, 445, image=self.inidcator_bg_image1, anchor="nw")

        self.grados_text = self.canvas.create_text(100, 470, text="Grados:", font=("Inter", 13), fill=self.text_color)
        self.torque_text = self.canvas.create_text(360, 470, text="Torque:", font=("Inter", 13), fill=self.text_color)

        self.gauge_bg_image = PhotoImage(file=relative_to_assets("GAUGE_BG0.png"))
        self.canvas.create_image(562, 135, image=self.gauge_bg_image, anchor="nw")

        for i in range(5):
            self.square_widget = tk.Label(self.canvas, text=str(i + 1), font=("Inter", 10), width=12, height=4, bd=0,
                                          highlightbackground="#D3D4D9",
                                          highlightthickness=1, bg="white")
            self.square_widget.place(x=583, y=398 - (i * 60))
            self.squares.append(self.square_widget)

        self.titleC_label = self.canvas.create_text(570, 120, text="Niveles", font=("Inter", 13), fill=self.text_color
                                                    , width=150, anchor="w")

        self.imagen_iniciar = PhotoImage(file=relative_to_assets("START_BTN0.png"))
        self.imagen_detener = PhotoImage(file=relative_to_assets("STOP_BTN0.png"))

        self.save_image = PhotoImage(file=relative_to_assets("SAVE_BTN0.png"))
        self.boton_save = tk.Button(
            self.canvas,
            image=self.save_image,
            state="normal",
            relief="flat",
            borderwidth=0,
            bg=self.used_color,
            command=self.save_boton)
        self.boton_save.place(x=720, y=550)
        self.boton_save.config(state="disabled")

        self.boton_toggle = tk.Button(
            self.canvas,
            image=self.imagen_iniciar,
            text="Iniciar",
            font=("Inter", 16),
            command=self.toggle_boton,
            state="normal",
            relief="flat",
            borderwidth=0,
            bg=self.used_color,
        )
        self.boton_toggle.place(x=734, y=410)
        self.boton_toggle.config(state="disabled")

        self.combobox.bind("<<ComboboxSelected>>", self.update_state)

        self.achieved_levels = [False] * 5

    def achieved_test(self, color):
        nivel_actual = self.combobox.get()
        if nivel_actual in self.datos:
            self.datos[nivel_actual].append({
                "Grados": self.grados,
                "Torque": self.torque,
                "Llegó": "Sí" if color == "#06D7A0" else "No"  # Dependiendo del color, sabes si llegó o no
            })
        self.highlight(color)
        self.combobox.config(state="readonly")
        self.turn_off_motor()

    def failed_test(self, color):
        nivel_actual = self.combobox.get()
        if nivel_actual in self.datos:
            self.datos[nivel_actual].append({
                "Grados": self.grados,
                "Torque": self.torque,
                "Llegó": "Sí" if color == "#06D7A0" else "No"  # Dependiendo del color, sabes si llegó o no
            })
        self.highlight(color)
        self.combobox.config(state="readonly")
        self.turn_off_motor()

    def apply_combox_changes(self):
        self.level = self.combobox.get()
        self.value = self.user_input.get()
        guide = (self.patient_data.get("Actividad", "No Registrado")).lower()
        key_id = guide[0:3]
        lvl4_max = int(read_input_from_json(f"{key_id}_nv4"))
        lvl5_max = int(read_input_from_json(f"{key_id}_nv5"))
        if self.level.startswith("Nivel "):
            self.level_num = int(self.level.split(" ")[1])

            if self.level_num == 4 and self.value.isdigit() and int(self.value) > lvl4_max:
                self.mensaje_label1.config(text=f"ERROR: Máx. {lvl4_max} N/m", fg="#00072d", bg="#FFFFFF")
                return
            elif self.level_num == 5 and self.value.isdigit() and int(self.value) > lvl5_max:
                self.mensaje_label1.config(text=f"ERROR: Máx. {lvl5_max} N/m", fg="#00072d", bg="#FFFFFF")
                return
            elif self.level_num in (4, 5) and ((not self.value.strip() or not self.value.isdigit())
                                               or int(self.value) == 0):
                self.mensaje_label1.config(text="ERROR: valor inválido", fg="#00072d", bg="#FFFFFF")
                return
            if not self.check_last_lvl(self.level_num):
                return

            self.mensaje_label1.config(text="Cambios aplicados", fg="#00072D")
            self.user_input.config(state="disabled")
            self.boton_toggle.config(state="normal")
            self.combobox.config(state="disabled")

        self.mensaje_label1.after(5000, lambda: self.mensaje_label1.config(text=""))

    def update_frame(self, position, torque):
        """Actualiza los datos de grados y torque."""
        self.grados = position
        self.torque = torque

        # Guardar los datos en self.datos
        nivel_actual = self.combobox.get()
        if nivel_actual in self.datos:
            self.datos[nivel_actual].append({
                "Grados": self.grados,  # Guardar los grados
                "Torque": self.torque,  # Guardar el torque
                "Llegó": "Sí"  # O "No", dependiendo de la lógica de tu aplicación
            })

    def save_boton(self):
        try:
            # Crear una lista para almacenar todos los datos
            datos_finales = []

            # Recorrer el diccionario y agregar los datos a la lista
            for nivel, registros in self.datos.items():
                for registro in registros:
                    datos_finales.append({
                        "Nivel": nivel,
                        "Grados": registro["Grados"],
                        "Torque": registro["Torque"],
                        "Llegó": registro["Llegó"]
                    })

            # Obtener la ruta de salida y el número de expediente desde patient_data
            expediente = self.patient_data.get("Expediente", "Sin expediente")
            output_folder = self.patient_data.get("output_folder", Path.cwd())

            if expediente and output_folder:
                # Crear el nombre del archivo con el número de expediente
                output_path = output_folder / f"{expediente}.xlsx"

                # Verificar si el archivo ya existe
                if output_path.exists():
                    # Abrir el archivo existente
                    wb = openpyxl.load_workbook(output_path)
                    # Verificar si ya existe una hoja llamada "Modo Manual"
                    if "Modo Manual" in wb.sheetnames:
                        # Si existe, eliminarla para evitar duplicados
                        wb.remove(wb["Modo Manual"])
                    # Crear una nueva hoja para el modo manual
                    ws_manual = wb.create_sheet("Modo Manual")
                else:
                    # Si el archivo no existe, crear uno nuevo
                    wb = openpyxl.Workbook()
                    # Crear la hoja para el modo manual
                    ws_manual = wb.active
                    ws_manual.title = "Modo Manual"

                # Escribir la tabla de información personal
                self.write_patient_info(ws_manual)

                # Escribir la tabla de resumen de pruebas de fuerza
                self.write_test_summary(ws_manual, datos_finales)

                # Crear la gráfica de barras
                chart = BarChart()
                chart.title = "Grados por Nivel"
                chart.x_axis.title = "Nivel"  # Eje X: Niveles
                chart.y_axis.title = "Grados"  # Eje Y: Grados

                # Referencias para los datos de la gráfica
                data = Reference(ws_manual, min_col=2, min_row=8,
                                 max_row=8 + len(datos_finales))  # Columna B (Grados)
                categories = Reference(ws_manual, min_col=1, min_row=8,
                                       max_row=8 + len(datos_finales))  # Columna A (Nivel)

                # Agregar los datos a la gráfica
                chart.add_data(data, titles_from_data=False)  # titles_from_data=False para evitar usar B8 como título
                chart.set_categories(categories)

                # Configurar las etiquetas de datos (valores arriba de las barras)
                chart.dLbls = None  # Deshabilitar las etiquetas de datos

                # Agregar la gráfica a la hoja
                ws_manual.add_chart(chart, "F7")  # Colocar la gráfica al lado de la tabla

                # Guardar el archivo Excel
                wb.save(output_path)

                # Mostrar mensaje de éxito
                messagebox.showinfo("Test Finalizado",
                                    f"El registro y test ha concluido. Su archivo ha sido guardado como {output_path}.")
            else:
                messagebox.showwarning("Error",
                                       "Por favor, ingrese un número de expediente válido y seleccione una carpeta de salida.")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al guardar el archivo: {e}")

        # Deshabilitar el botón de guardar y resetear la interfaz
        self.boton_save.config(state="disabled")
        self.user_input.config(state="disabled")
        self.squares[4].config(bg="#FFFFFF")
        self.squares[3].config(bg="#FFFFFF")
        self.squares[2].config(bg="#FFFFFF")
        self.squares[1].config(bg="#FFFFFF")
        self.squares[0].config(bg="#FFFFFF")
        self.achieved_levels[:] = [False] * len(self.achieved_levels)
        self.datos = {
            "Nivel 1": [],
            "Nivel 2": [],
            "Nivel 3": [],
            "Nivel 4": [],
            "Nivel 5": []
        }

    def write_patient_info(self, ws):
        """Escribe la información del paciente en la hoja de Excel."""
        ws['A1'] = "Información Personal"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A1:E1')  # Combinar celdas para el header (ahora son 5 columnas)

        # Subheaders de la tabla de información personal
        ws['A2'] = "Nombre"
        ws['B2'] = "Edad"
        ws['C2'] = "Sexo"
        ws['D2'] = "Actividad Física"
        ws['E2'] = "Fecha de prueba"
        for cell in ws['A2:E2'][0]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos de la tabla de información personal
        ws['A3'] = self.patient_data.get("Nombre", "No registrado")
        ws['B3'] = self.patient_data.get("Edad", "No registrado")
        ws['C3'] = self.patient_data.get("Sexo", "No registrado")
        ws['D3'] = self.patient_data.get("Actividad", "No registrado")
        ws['E3'] = self.patient_data.get("Fecha de prueba",
                                         "No registrado")  # Puedes agregar un campo para la fecha si es necesario
        for cell in ws['A3:E3'][0]:
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def write_test_summary(self, ws, datos_finales):
        """Escribe el resumen de las pruebas de fuerza en la hoja de Excel."""
        # Espacio entre las tablas
        ws['A5'] = ""  # Espacio vacío

        # Escribir la tabla de resumen de pruebas de fuerza
        ws['A6'] = "Resumen de Pruebas de Fuerza"
        ws['A6'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A6'].fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        ws['A6'].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells('A6:D6')  # Combinar celdas para el header

        # Subheaders de la tabla de resumen de pruebas de fuerza
        ws['A7'] = "Nivel"
        ws['B7'] = "Grados"
        ws['C7'] = "Torque"
        ws['D7'] = "Llegó"
        for cell in ws['A7:D7'][0]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Datos de la tabla de resumen de pruebas de fuerza
        row_index = 8
        for dato in datos_finales:
            ws[f'A{row_index}'] = dato["Nivel"]
            ws[f'B{row_index}'] = dato["Grados"]  # Asegúrate de que los grados se escriban aquí
            ws[f'C{row_index}'] = dato["Torque"]
            ws[f'D{row_index}'] = dato["Llegó"]
            for cell in ws[f'A{row_index}:D{row_index}'][0]:
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_index += 1

    def start_animation(self):
        self.level1 = self.combobox.get()
        if self.level1.startswith("Nivel "):
            self.level1_num = int(self.level1.split(" ")[1])
            if 1 <= self.level1_num <= 5:
                self.active_animation = True
                self.blink_state = True
                self.blinking(self.squares[self.level1_num - 1])
                self.combobox.config(state="disabled")
                self.achieved_button.config(state="normal")
                self.failed_button.config(state="normal")
                self.user_input.config(state="disabled")
                self.boton_save.config(state="disabled")
                self.boton_toggle.config(text="Detener", image=self.imagen_detener)

    def blinking(self, square):
        if self.active_animation:
            color = "#60AFFF" if self.blink_state else "white"
            square.config(bg=color)
            self.blink_state = not self.blink_state
            square.after(500, lambda: self.blinking(square))

    def stop_animation(self):
        self.level = self.combobox.get()
        self.active_animation = False
        self.achieved_button.config(state="disabled")
        self.failed_button.config(state="disabled")
        self.combobox.config(state="normal")
        self.boton_toggle.config(text="Iniciar", image=self.imagen_iniciar, state="disabled")
        self.boton_save.config(state="normal")

        if self.level.startswith("Nivel "):
            self.nivel_num = int(self.level.split(" ")[1])
            if self.nivel_num > 3:
                self.user_input.config(state="normal")

    def highlight(self, color):
        self.level = self.combobox.get()

        if self.level.startswith("Nivel "):
            self.level_num = int(self.level.split(" ")[1])
            if 1 <= self.level_num <= 3:
                self.stop_animation()
                self.squares[self.level_num - 1].config(bg=color)
                self.achieved_levels[self.level_num - 1] = True
                self.user_input.config(state="disabled")
            else:
                self.stop_animation()
                self.squares[self.level_num - 1].config(bg=color)
                self.achieved_levels[self.level_num - 1] = True
                self.user_input.config(state="normal")

    def init_widgets(self):
        # Mostrar información del paciente
        nombre_paciente = self.patient_data.get("Nombre", "No registrado")
        self.nombre_title = self.canvas.create_text(50, 120, text=f"Pierna de: {nombre_paciente}",
                                                    fill=self.text_color, font=("Inter", 13), width=400, anchor="w")

        self.return_image = PhotoImage(file=relative_to_assets("GO_BACK_BTN1.png"))
        self.return_btn = tk.Button(self.canvas, state="normal", relief="flat",
                                    borderwidth=0,
                                    bg=self.used_color,
                                    image=self.return_image,
                                    command=lambda: self.controller.switch_frame(AppInterface1))
        self.return_btn.place(x=865, y=645)

    def on_closing(self):
        self.turn_off_motor()
        time.sleep(0.05)
        self.disconnect_arduino()
        self.root.destroy()


"""
############################################## SEXTA VENTANA ###########################################################
Función que controla la imagen de la pierna

########################################################################################################################
"""


class LegAnimation:
    def __init__(self, canvas, image_folder_path):
        self.canvas = canvas
        self.image_folder_path = image_folder_path

        # Load PNG frames from the folder
        self.frames = self.load_png_frames(image_folder_path)

        # Create a label to display the image frame
        self.label = tk.Label(self.canvas, bg="white", highlightthickness=0)
        self.label.place(x=46, y=137)  # Place the label

        self.text_id1 = self.canvas.create_text(
            240, 470,  # Coordinates (x, y)
            text="0.0°",  # Text content
            font=("Inter", 13),  # Font and size
            fill="black"  # Text color
        )

        self.text_id2 = self.canvas.create_text(
            490, 470,  # Coordinates (x, y)
            text="0.0",  # Text content
            font=("Inter", 13),  # Font and size
            fill="black"  # Text color
        )

        # Initialize with the first frame
        self.update_frame(0, 0)  # Start with position 0

    def load_png_frames(self, image_folder_path):
        frames = []
        # Sort files to ensure correct order
        files = sorted(Path(image_folder_path).iterdir(), key=lambda x: int(x.stem.split("_")[-1]))
        for file in files:
            if file.is_file() and file.suffix.lower() == ".png":
                image = Image.open(file)
                frames.append(image)
        return frames

    def input_to_frame(self, position, total_frames, degrees_per_frame=2.3):
        frame_index = int(position / degrees_per_frame)
        # Ensure the frame index is within bounds
        frame_index = max(0, min(frame_index, total_frames - 1))
        return frame_index

    def update_frame(self, position, torque):
        """Update the displayed frame based on position value."""
        frame_index = self.input_to_frame(position, len(self.frames))
        frame = self.frames[frame_index]
        frame_image = ImageTk.PhotoImage(image=frame)
        self.label.config(image=frame_image)
        self.label.image = frame_image  # Keep a reference to avoid garbage collection

        # Update the text with the current value
        self.canvas.itemconfig(self.text_id1, text=f"{position:.1f}°")
        self.canvas.itemconfig(self.text_id2, text=f"{torque:.1f}  N/m")


if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("630x550")
    controller = Controller(root)  # Create the controller
    controller.switch_frame(AppInterface0)  # Start with AppInterface0
    root.mainloop()
