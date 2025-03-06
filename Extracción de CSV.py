import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference


def guardar_registro():
    nombre = entry_nombre.get().strip()
    edad = entry_edad.get().strip()
    sexo = combo_sexo.get().strip()
    expediente = entry_expediente.get().strip()

    if not (nombre and edad and sexo and expediente):
        messagebox.showwarning("Error", "Todos los campos son obligatorios")
        return

    archivo = f"{expediente}.xlsx"

    # Verificar si el archivo existe o crear uno nuevo
    if os.path.exists(archivo):
        wb = load_workbook(archivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Nombre", "Edad", "Sexo", "Expediente"])  # Encabezados generales

    # Aplicar formato a encabezados
    header_fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
    header_font = Font(bold=True)

    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).font = header_font

    # Agregar los datos personales solo en la fila 1 si no están ya
    if ws.max_row == 1:
        ws.append([nombre, edad, sexo, expediente])

    # Agregar encabezados para la tabla de tests si no existen
    if ws.max_row < 4:
        ws.append([])
        ws.append(["Test", "Nivel 1", "Nivel 2", "Nivel 3"])  # Encabezados de la tabla

        for col in range(1, 5):
            ws.cell(row=ws.max_row, column=col).fill = header_fill
            ws.cell(row=ws.max_row, column=col).font = header_font

    # Agregar datos de los tests
    fila_inicio = ws.max_row + 1
    for i in range(5):
        fila_test = [f"Test {i + 1}"] + [float(entries_niveles[i][j].get().strip() or 0) for j in range(3)]
        ws.append(fila_test)

    # Crear gráfica de barras
    chart = BarChart()
    chart.title = "Resultados de Tests"
    chart.x_axis.title = "Tests"
    chart.y_axis.title = "Valores"

    datos = Reference(ws, min_col=2, max_col=4, min_row=fila_inicio - 1, max_row=fila_inicio + 4)
    categorias = Reference(ws, min_col=1, min_row=fila_inicio, max_row=fila_inicio + 4)
    chart.add_data(datos, titles_from_data=True)
    chart.set_categories(categorias)

    ws.add_chart(chart, "F5")  # Ubicación de la gráfica en la hoja

    # Guardar archivo
    wb.save(archivo)

    messagebox.showinfo("Éxito", f"Registro guardado en {archivo}")
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_expediente.delete(0, tk.END)
    combo_sexo.set("")
    for row in entries_niveles:
        for entry in row:
            entry.delete(0, tk.END)


# Crear ventana principal
root = tk.Tk()
root.title("Registro de Datos")
root.geometry("500x600")

# Etiquetas y entradas
ttk.Label(root, text="Nombre:").pack(pady=5)
entry_nombre = ttk.Entry(root)
entry_nombre.pack()

ttk.Label(root, text="Edad:").pack(pady=5)
entry_edad = ttk.Entry(root)
entry_edad.pack()

ttk.Label(root, text="Sexo:").pack(pady=5)
combo_sexo = ttk.Combobox(root, values=["Masculino", "Femenino", "Otro"])
combo_sexo.pack()

ttk.Label(root, text="N° Expediente:").pack(pady=5)
entry_expediente = ttk.Entry(root)
entry_expediente.pack()

# Tabla de Test y Niveles
ttk.Label(root, text="Resultados de Tests").pack(pady=10)
table_frame = ttk.Frame(root)
table_frame.pack()

headers = ["", "Nivel 1", "Nivel 2", "Nivel 3"]
for col, text in enumerate(headers):
    ttk.Label(table_frame, text=text, borderwidth=2, relief="ridge", width=12).grid(row=0, column=col)

entries_niveles = []
for i in range(1, 6):
    row_entries = []
    ttk.Label(table_frame, text=f"Test {i}", borderwidth=2, relief="ridge", width=12).grid(row=i, column=0)
    for j in range(1, 4):
        entry = ttk.Entry(table_frame, width=10)
        entry.grid(row=i, column=j)
        row_entries.append(entry)
    entries_niveles.append(row_entries)

# Botón de guardar
btn_guardar = ttk.Button(root, text="Guardar Registro", command=guardar_registro)
btn_guardar.pack(pady=10)

root.mainloop()
